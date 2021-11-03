#! python3.9
"""Pull the latest treasury rate data from the Federal Reserve Economic Database (FRED) and save to Excel file."""

import json
import urllib.request as url_request

import openpyxl

# get a free API key from https://fredaccount.stlouisfed.org/apikeys
API_KEY = "XXXXXXXXXXXXXXXXXXXXXX"

# list of treasury securities to pull rates for
TREASURIES = [
    "DGS1MO",
    "DGS3MO",
    "DGS6MO",
    "DGS1",
    "DGS2",
    "DGS3",
    "DGS5",
    "DGS7",
    "DGS10",
    "DGS20",
    "DGS30",
]

urlopen = url_request.urlopen


def get_rates():
    
    """
    Gets the most recent interest rate of each treasury security in list.
    Returns a nested dictionary of treasury/interest rate pairs.
    """

    memo = {}
    data_date = None
    for treasury in TREASURIES:
        URL = f"https://api.stlouisfed.org/fred/series/observations?series_id={treasury}&observation_start=2021-10-01&sort_order=asc&api_key={API_KEY}&file_type=json"

        response = urlopen(URL)
        historical_rates = json.loads(response.read())

        latest_rate = historical_rates["observations"][-1]  # only pull the last (most recent) treasury rate
        obs_date = latest_rate["date"]
        memo[treasury] = latest_rate["value"]

        """
        this block is to ensure that each loop's data is for the same date.
        if for some reason one treasury's latest interest rate is for a different date than others, the data
        won't be able to output to Excel as needed. To this point, I haven't had this error occur.
        """
        if not data_date:
            data_date = obs_date
        else:
            if obs_date != data_date:
                output = "Error - Multiple Dates."
                return output
    memo["Date"] = data_date  # store the date of the gathered data. This will be used in Excel output

    return memo


def convert_excel(treasuries):
    """Convert treasury rate data into an Excel spreadsheet."""

    outfile = "output/treasuries_" + treasuries["Date"] + ".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = "Date"
    ws.cell(row=2, column=1).value = treasuries["Date"]

    for idx, header in enumerate(TREASURIES):
        ws.cell(row=1, column=idx + 2).value = header
        ws.cell(row=2, column=idx + 2).value = treasuries[header]

    wb.save(outfile)


def main():
    latest_treasuries = get_rates()
    convert_excel(latest_treasuries)


if __name__ == "__main__":
    main()
